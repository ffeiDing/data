import openpyxl
import json
import re
import zhconv
from pypinyin import lazy_pinyin

def hans_2_hant(hans_str: str):
    '''
    Function: 将 hans_str 由简体转化为繁体
    '''
    return zhconv.convert(hans_str, 'zh-hant')

def hant_2_hans(hant_str: str):
    '''
    Function: 将 hant_str 由繁体转化为简体
    '''
    return zhconv.convert(hant_str, 'zh-hans')

workbook = openpyxl.load_workbook("data/rubbing.xlsx")
# workbook = openpyxl.load_workbook("data/METADATA.xlsx")
shenames = workbook.sheetnames
# print(shenames)

worksheet = workbook["Result 1"]
# print(worksheet) 

rows = worksheet.max_row
columns = worksheet.max_column
# print(rows, columns) #3

data = {}

first_row = True
idx = 0

for row in worksheet.rows: 
    if first_row:
        first_row = False
        continue
    else:
        xml = row[1].value
        res = re.findall(r"title titleAttribute=\"(.*?)\"", xml)
        # res = re.findall(r"\<gravureDesc\>(.*?)\<\/gravureDesc\>", xml)

        # res = re.findall(r"historyattribute=\"\"", xml)
        # if len(res) == 0:
        #     continue

        # res = re.findall(r"\<calis:titlevalue\>太白酒(.+?)\<\/calis:titlevalue\>", xml)
        # res = re.findall(r"\<calis:title titleattribute=\"正題名及説明\"\>(.+?)\<\/calis:title\>", xml)
        # if len(res) > 0:
        #     print(row[0].value, xml)

        if len(res) > 1:
            idx += 1
        if len(res) > 0:
            for str in res:
                # str = str.strip()
                # str = hant_2_hans(str)
                # str = ''.join(lazy_pinyin(str))
                data.setdefault(str, 0)
                data[str] += 1

for key in data.keys():
    print(key, data[key])

print(idx)
        
