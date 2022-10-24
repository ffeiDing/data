import openpyxl
import json
import re
import zhconv
from pypinyin import lazy_pinyin

data = {}

workbook = openpyxl.load_workbook("data/a_service_ap_log1.xlsx")
# workbook = openpyxl.load_workbook("data/METADATA.xlsx")
shenames = workbook.sheetnames
# print(shenames)

worksheet = workbook["Result 1"]
# print(worksheet) 

rows = worksheet.max_row
columns = worksheet.max_column
print(rows, columns) #3

first_row = True
idx = 0
pre_time = "-1"

for row in worksheet.rows: 
    if first_row:
        first_row = False
        continue
    else:
        ap_id = row[1].value
        num = row[2].value
        time = row[3].value
        if pre_time == "-1":
            data.setdefault(time, {})
            data[time][ap_id] = num
            pre_time = time
        else:
            minute_pre = int(pre_time.split(":")[-2])
            minute_cur = int(time.split(":")[-2])
            # print(minute_cur, minute_pre)
            if minute_cur == minute_pre or minute_cur-minute_pre <= 2 or (minute_pre == 59 and minute_cur <= 1) or (minute_pre == 58 and minute_cur == 0):
                data[pre_time][ap_id] = num
            else:
                data.setdefault(time, {})
                data[time][ap_id] = num
                pre_time = time


worksheet = workbook["Result 2"]
# print(worksheet) 

rows = worksheet.max_row
columns = worksheet.max_column
print(rows, columns) #3

first_row = True

for row in worksheet.rows: 
    if first_row:
        first_row = False
        continue
    else:
        ap_id = row[1].value
        num = row[2].value
        time = row[3].value
        if pre_time == "-1":
            data.setdefault(time, {})
            data[time][ap_id] = num
            pre_time = time
        else:
            minute_pre = int(pre_time.split(":")[-2])
            minute_cur = int(time.split(":")[-2])
            # print(minute_cur, minute_pre)
            if minute_cur == minute_pre or (minute_cur > minute_pre and minute_cur-minute_pre <= 2) or (minute_pre == 59 and minute_cur <= 1) or (minute_pre == 58 and minute_cur == 0):
                data[pre_time][ap_id] = num
            else:
                data.setdefault(time, {})
                data[time][ap_id] = num
                pre_time = time
# print(data)


data_sum = {}
for time in data.keys():
    for ap_id in data[time].keys():
        data_sum.setdefault(time, 0)
        data_sum[time] += data[time][ap_id]

import csv

def load_dic(name):
    with open(name, 'rb') as f:
        return pickle.load(f)

csvFile=open("data/ap.csv",'w',newline='')
writer=csv.writer(csvFile)
writer.writerow(("time", "num"))


for time, num in data_sum.items():
    writer.writerow((time, num))

csvFile.close()
