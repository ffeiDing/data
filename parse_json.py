# 先清洗数据，仅留下jsonb的action为checkedout的条目

import openpyxl
import json

workbook = openpyxl.load_workbook("data/user50000+.xlsx")
shenames = workbook.sheetnames
# print(shenames)

worksheet = workbook["user50000+"]
# print(worksheet) 

rows = worksheet.max_row
columns = worksheet.max_column
# print(rows, columns) #46854 73


cn_attr = ["出版年", "作者", "版本", "图书装帧", "语种", "卷", "页码", "价格", "系列", "出版社", "ISBN10", "ISBN13", "币种"]
en_attr = ["publish_year", "author", "edition", "binding", "language", "volume", "pages", "price", "series", "publish", "ISBN10", "ISBN13", "currency"]
clean_rows = []
first_row = True

clean_worksheet = workbook.create_sheet() 
clean_worksheet.title = "Parse_Result"

row_idx = 0
for row in worksheet.rows: 
    clean_row = []
    for cell_idx in range(len(row)):
        if cell_idx == 14 or cell_idx == 17:
            continue
        cell = row[cell_idx]
        clean_row.append(cell.value)
    if first_row:
        first_row = False
        for idx in range(len(cn_attr)):
            clean_row.append(cn_attr[idx]+"("+en_attr[idx]+")")
    else:
        jsonb = json.loads(row[17].value)["attrValueMap"]
        for idx in range(len(cn_attr)):
            if en_attr[idx] in jsonb:
                if "value" in jsonb[en_attr[idx]]:
                    attr_value = jsonb[en_attr[idx]]["value"]
            else:
                attr_value = ""
            clean_row.append(attr_value)
    print(clean_row)
    for col_idx in range(len(clean_row)):
        clean_worksheet.cell(row_idx+1, col_idx+1, clean_row[col_idx])
    row_idx = row_idx + 1

workbook.save(filename='data/processed_50000+.xlsx')