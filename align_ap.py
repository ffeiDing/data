import openpyxl
import json
import re
import zhconv
from pypinyin import lazy_pinyin
import csv
zhaji = []
zhaji_time = []
ap = []
ap_time = []

workbook = openpyxl.load_workbook("data/ap.xlsx")
# workbook = openpyxl.load_workbook("data/METADATA.xlsx")
shenames = workbook.sheetnames
# print(shenames)

worksheet = workbook["ap"]
# print(worksheet) 

rows = worksheet.max_row
columns = worksheet.max_column
print(rows, columns) #3

first_row = True
for row in worksheet.rows: 
    if first_row:
        first_row = False
        continue
    time = str(row[0].value)
    num = row[1].value    
    ap.append(num)
    # year = int(time.split("-")[0])
    # month = int(time.split("-")[1])
    # day = int(time.split(" ")[0].split("-")[-1])
    # ap_time.append(str(year)+"/"+str(month)+"/"+str(day)+" "+time.split(" ")[-1].split(":")[0])
    ap_time.append(time)

workbook = openpyxl.load_workbook("data/a_service_reader_in_library_log.xlsx")
# workbook = openpyxl.load_workbook("data/METADATA.xlsx")
shenames = workbook.sheetnames
# print(shenames)

worksheet = workbook["Result 1"]
# print(worksheet) 

rows = worksheet.max_row
columns = worksheet.max_column
print(rows, columns) #3

first_row = True
for row in worksheet.rows: 
    if first_row:
        first_row = False
        continue

    time = str(row[1].value)
    num = row[0].value
    zhaji_time.append(time)
    zhaji.append(num)


csvFile=open("data/align.csv",'w',newline='')
writer=csv.writer(csvFile)
writer.writerow(("time", "ap", "zhaji"))

data = {}
i = 0
j = 0
while i < len(zhaji_time):
    tmp1 = zhaji_time[i]
    day1 = tmp1.split(" ")[0]
    while j < len(ap_time):
        tmp2 = ap_time[j]
        day2 = tmp2.split(" ")[0]
        if day1 < day2:
            i += 1
            break
        if day2 < day1:
            j += 1
            continue
        hour1 = int(tmp1.split(" ")[-1].split(":")[0])
        hour2 = int(tmp2.split(" ")[-1].split(":")[0])
        minute1 = int(tmp1.split(" ")[-1].split(":")[1])
        minute2 = int(tmp2.split(" ")[-1].split(":")[1])
        print(i, j, hour1, hour2, minute1, minute2)
        if (hour1 + 1 < hour2) or ((hour1 + 1 == hour2) and (minute1 < minute2 or (60+minute2-minute1) > 10)):
            i += 1
            break
        if (hour2 + 1 < hour1) or ((hour2 + 1 == hour1) and (minute2 < minute1 or (60+minute1-minute2) > 10)):
            j += 1
            continue
        writer.writerow((tmp1, ap[j], zhaji[i]))
        i += 1
        j += 1
        break

csvFile.close()

